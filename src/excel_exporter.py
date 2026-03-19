from __future__ import annotations

from typing import List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def _write_heading(ws: Worksheet, row: int, text: str) -> int:
    ws.cell(row=row, column=1, value=text)
    return row + 1


def _write_table(ws: Worksheet, start_row: int, headers: List[str], rows: List[List[str]]) -> int:
    row = start_row

    # headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)
    row += 1

    for r in rows:
        # skip completely empty rows
        if all(str(cell).strip() == "" for cell in r):
            continue

        for col_idx, val in enumerate(r, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        row += 1

    return row


def export_all_tables_to_excel(
    output_path: str,
    table1,
    table2,
    table3,
    table4,
    table5,
    table6,
    table7,
    table8,
):
    wb = Workbook()
    ws = wb.active

    country = "Unknown"
    fmv = "Low"

    for item in table1.items:
        if item.category.lower() == "country" and item.details:
            country = item.details
        elif item.category.lower() == "fmv" and item.details:
            fmv = item.details

    ws.title = f"{country} - {fmv}"

    row = 1

    # 1. Site Budget Estimation
    row = _write_heading(ws, row, "1. Site Budget Estimation")
    headers = ["Category", "Details"]
    rows = [[x.category, x.details] for x in table1.items]
    row = _write_table(ws, row, headers, rows)
    row += 2

    # 2. Study Details Table
    row = _write_heading(ws, row, "2. Study Details Table")
    headers = ["Field", "Value"]

    def _stringify(v):
        if v is None:
            return ""
        if hasattr(v, "value"):
            return "" if v.value is None else str(v.value)
        if isinstance(v, dict) and "value" in v:
            return "" if v["value"] is None else str(v["value"])
        return str(v)

    rows = []
    preferred_order = [
        "document_date",
        "protocol_number",
        "indication",
        "phase",
        "protocol_title",
    ]

    t2 = table2.model_dump()
    for key in preferred_order:
        if key in t2 and key != "pdf_name":
            rows.append([key, _stringify(t2.get(key))])

    for k, v in t2.items():
        if k not in preferred_order and k != "pdf_name":
            rows.append([k, _stringify(v)])

    row = _write_table(ws, row, headers, rows)
    row += 2

    # 3. Visit Details Table
    row = _write_heading(ws, row, "3. Visit Details Table")
    headers = ["Visit Title", "Visit ID"]
    rows = [[x.visit_title, x.visit_id] for x in table3.visits]
    row = _write_table(ws, row, headers, rows)
    row += 2

    visit_ids = [v.visit_id for v in table3.visits]

    # 4. List of Procedures
    row = _write_heading(ws, row, "4. List of Procedures")
    headers = ["Procedure", "Code", "Unit Basis", "Budget"] + visit_ids

    rows = []
    for x in table4.procedures:
        row_data = [
            getattr(x, "procedure", ""),
            getattr(x, "code", ""),
            getattr(x, "unit_basis", ""),
            getattr(x, "budget", ""),
        ]
        visit_values = getattr(x, "visit_values", {}) or {}
        for vid in visit_ids:
            row_data.append(visit_values.get(vid, ""))
        rows.append(row_data)

    row = _write_table(ws, row, headers, rows)
    row += 2

    # 5. Non Procedures
    # =========================
    row = _write_heading(ws, row, "5. Non Procedures")

    headers = ["Non Procedure Item", "Code", "Unit Basis", "Budget"] + visit_ids

    rows = []
    for x in table5.items:
        name = getattr(x, "non_procedure_item", "").strip()

        if not name:
            continue  # 🚫 skip empty rows

        row_data = [
            name,
            getattr(x, "code", ""),
            getattr(x, "unit_basis", ""),
            getattr(x, "budget", ""),
        ]

        visit_values = getattr(x, "visit_values", {}) or {}
        for vid in visit_ids:
            row_data.append(visit_values.get(vid, ""))

        rows.append(row_data)

    row = _write_table(ws, row, headers, rows)
    row += 2

    # 6. Site Fees
    row = _write_heading(ws, row, "6. Site Fees")
    headers = ["Site Fee Description", "Code", "Unit Basis", "Unit Cost"]

    rows = []
    for x in table6.items:
        description = getattr(x, "site_fee_description", None)
        if description is None:
            description = getattr(x, "description", "")
        rows.append([
            description,
            getattr(x, "code", ""),
            getattr(x, "unit_basis", ""),
            getattr(x, "unit_cost", ""),
        ])

    row = _write_table(ws, row, headers, rows)
    row += 2

    # 7. Conditional Procedures
    row = _write_heading(ws, row, "7. Conditional Procedures")
    headers = [
        "Conditional Procedures",
        "Code",
        "Unit Basis",
        "Unit Cost",
        "Overhead",
        "Unit cost incl. Overhead",
    ]

    rows = []
    for x in table7.items:
        proc = getattr(x, "conditional_procedure", None)
        if proc is None:
            proc = getattr(x, "procedure", "")
        rows.append([
            proc,
            getattr(x, "code", ""),
            getattr(x, "unit_basis", ""),
            getattr(x, "unit_cost", ""),
            getattr(x, "overhead", ""),
            getattr(x, "unit_cost_incl_overhead", ""),
        ])

    row = _write_table(ws, row, headers, rows)
    row += 2

    # 8. Patient Costs
    row = _write_heading(ws, row, "8. Patient Costs")
    headers = ["Patient Cost Item", "Code", "Unit Basis", "Unit Cost"]

    rows = []
    for x in table8.items:
        rows.append([
            getattr(x, "patient_cost_item", ""),
            getattr(x, "code", ""),
            getattr(x, "unit_basis", ""),
            getattr(x, "unit_cost", ""),
        ])

    _write_table(ws, row, headers, rows)

    wb.save(output_path)