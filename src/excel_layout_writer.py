from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


SECTION_TITLES = [
    "1. Site Budget Estimation",
    "2. Study Details Table",
    "3. Visit Details Table",
    "4. List of Procedures",
    "5. Non Procedures",
    "6. Site Fees",
    "7. Conditional Procedures",
    "8. Patient Costs",
]

HEADER_FILL = "A9D0F5"  # light blue close to the template


def _normalize_title(text: str | None) -> str:
    return (text or "").replace(",", ".").strip().lower()


def _find_row_by_title(ws: Worksheet, title: str) -> int:
    target = _normalize_title(title)
    for row in range(1, ws.max_row + 1):
        if _normalize_title(str(ws.cell(row, 1).value)) == target:
            return row
    raise ValueError(f"Could not find section title: {title}")


def _find_all_section_rows(ws: Worksheet) -> Dict[str, int]:
    return {title: _find_row_by_title(ws, title) for title in SECTION_TITLES}


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def _copy_row_format(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        _copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def _write_row_values(ws: Worksheet, row_idx: int, values: Sequence[Any], max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row_idx, col).value = None
    for offset, value in enumerate(values, start=1):
        ws.cell(row_idx, offset).value = value


def _set_header_fill(ws: Worksheet, row_idx: int, columns: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    for col in range(1, columns + 1):
        ws.cell(row_idx, col).fill = fill


def _table1_rows(table1: Mapping[str, Any]) -> List[List[Any]]:
    rows = []
    for item in table1["items"]:
        rows.append([item["category"], item["details"]])
    return rows


def _table2_rows(table2: Mapping[str, Any]) -> List[List[Any]]:
    def _val(v: Any) -> Any:
        if isinstance(v, dict):
            return v.get("value") or ""
        return v or ""

    return [
        ["Protocol date:", _val(table2.get("document_date"))],
        ["Protocol number:", _val(table2.get("protocol_number"))],
        ["Indication:", _val(table2.get("indication"))],
        ["Phase:", _val(table2.get("phase"))],
        ["Protocol Title:", _val(table2.get("protocol_title"))],
    ]


def _table3_rows(table3: Mapping[str, Any]) -> List[List[Any]]:
    return [[v["visit_title"], v["visit_id"]] for v in table3["visits"]]


def _table4_rows(table4: Mapping[str, Any]) -> List[List[Any]]:
    visit_cols = table4["visit_columns"]
    rows: List[List[Any]] = []
    for item in table4["procedures"]:
        vals = [
            item.get("procedure", ""),
            item.get("code", ""),
            item.get("unit_basis", ""),
            item.get("budget", ""),
        ]
        vals.extend(item.get("visit_values", {}).get(col, "") for col in visit_cols)
        rows.append(vals)
    return rows


def _table5_rows(table5: Mapping[str, Any]) -> List[List[Any]]:
    visit_cols = table5["visit_columns"]
    rows: List[List[Any]] = []
    for item in table5["items"]:
        vals = [
            item.get("non_procedure_item", ""),
            item.get("code", ""),
            item.get("unit_basis", ""),
            item.get("budget", ""),
        ]
        vals.extend(item.get("visit_values", {}).get(col, "") for col in visit_cols)
        rows.append(vals)
    return rows


def _table6_rows(table6: Mapping[str, Any]) -> List[List[Any]]:
    return [
        [i.get("site_fee_description", ""), i.get("code", ""), i.get("unit_basis", ""), i.get("unit_cost", "")]
        for i in table6["items"]
    ]


def _table7_rows(table7: Mapping[str, Any]) -> List[List[Any]]:
    return [
        [
            i.get("conditional_procedure", ""),
            i.get("code", ""),
            i.get("unit_basis", ""),
            i.get("unit_cost", ""),
            i.get("overhead", ""),
            i.get("unit_cost_incl_overhead", ""),
        ]
        for i in table7["items"]
    ]


def _table8_rows(table8: Mapping[str, Any]) -> List[List[Any]]:
    return [
        [i.get("patient_cost_item", ""), i.get("code", ""), i.get("unit_basis", ""), i.get("unit_cost", "")]
        for i in table8["items"]
    ]


def _write_section(
    ws: Worksheet,
    title_row: int,
    heading: str,
    header: Sequence[Any],
    data_rows: List[List[Any]],
    total_cols: int,
    source_header_row: int,
    source_data_row: int,
) -> int:
    """Write one section starting at title_row. Returns next free row after two blank rows."""
    # Title
    _copy_row_format(ws, title_row, title_row, total_cols)
    ws.cell(title_row, 1).value = heading

    # Header row
    header_row = title_row + 1
    _copy_row_format(ws, source_header_row, header_row, total_cols)
    _write_row_values(ws, header_row, header, total_cols)
    _set_header_fill(ws, header_row, len(header))

    # Data rows
    data_start = header_row + 1
    for idx, row_vals in enumerate(data_rows):
        row_idx = data_start + idx
        _copy_row_format(ws, source_data_row, row_idx, total_cols)
        _write_row_values(ws, row_idx, row_vals, total_cols)

    # Two blank rows after each section
    blank1 = data_start + len(data_rows)
    blank2 = blank1 + 1
    _copy_row_format(ws, source_data_row, blank1, total_cols)
    _copy_row_format(ws, source_data_row, blank2, total_cols)
    _write_row_values(ws, blank1, [], total_cols)
    _write_row_values(ws, blank2, [], total_cols)
    return blank2 + 1


def build_budget_workbook(
    template_path: str | Path,
    output_path: str | Path,
    table1: Mapping[str, Any],
    table2: Mapping[str, Any],
    table3: Mapping[str, Any],
    table4: Mapping[str, Any],
    table5: Mapping[str, Any],
    table6: Mapping[str, Any],
    table7: Mapping[str, Any],
    table8: Mapping[str, Any],
    sheet_name: str = "Sheet1",
) -> Path:
    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    total_cols = max(ws.max_column, 16)
    sec_rows = _find_all_section_rows(ws)

    # Clear sheet below row 1 by rebuilding from top section down.
    # We keep column widths and overall workbook formatting intact.
    current_row = sec_rows[SECTION_TITLES[0]]

    current_row = _write_section(
        ws,
        current_row,
        SECTION_TITLES[0],
        ["Category", "Details"],
        _table1_rows(table1),
        total_cols,
        source_header_row=2,
        source_data_row=3,
    )
    current_row = _write_section(
        ws,
        current_row,
        SECTION_TITLES[1],
        ["Study Details", "Details"],
        _table2_rows(table2),
        total_cols,
        source_header_row=7,
        source_data_row=8,
    )
    current_row = _write_section(
        ws,
        current_row,
        SECTION_TITLES[2],
        ["Visit Title", "Visit ID"],
        _table3_rows(table3),
        total_cols,
        source_header_row=15,
        source_data_row=16,
    )
    current_row = _write_section(
        ws,
        current_row,
        SECTION_TITLES[3],
        ["Procedure", "Code", "Unit Basis", "Budget", *table4["visit_columns"]],
        _table4_rows(table4),
        total_cols,
        source_header_row=20,
        source_data_row=21,
    )
    current_row = _write_section(
        ws,
        current_row,
        SECTION_TITLES[4],
        ["Non Procedure Item", "Code", "Unit Basis", "Budget", *table5["visit_columns"]],
        _table5_rows(table5),
        total_cols,
        source_header_row=27,
        source_data_row=28,
    )
    current_row = _write_section(
        ws,
        current_row,
        SECTION_TITLES[5],
        ["Site Fee Description", "Code", "Unit Basis", "Unit Cost"],
        _table6_rows(table6),
        total_cols,
        source_header_row=45,
        source_data_row=46,
    )
    current_row = _write_section(
        ws,
        current_row,
        SECTION_TITLES[6],
        ["Conditional Procedures", "Code", "Unit Basis", "Unit Cost", "Overhead", "Unit cost incl. Overhead"],
        _table7_rows(table7),
        total_cols,
        source_header_row=74,
        source_data_row=75,
    )
    current_row = _write_section(
        ws,
        current_row,
        SECTION_TITLES[7],
        ["Patient Cost Item", "Code", "Unit Basis", "Unit Cost"],
        _table8_rows(table8),
        total_cols,
        source_header_row=79,
        source_data_row=80,
    )

    # Clear any residual old content below the final section to keep workbook clean.
    for r in range(current_row, ws.max_row + 1):
        for c in range(1, total_cols + 1):
            ws.cell(r, c).value = None

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
